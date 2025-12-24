from flask import Flask, request, jsonify, send_from_directory
from openpyxl import load_workbook, Workbook
from flask_cors import CORS
import os
import pdfplumber
import re
import uuid

app = Flask(__name__)
CORS(app)

# --- CONFIGURAÇÕES ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_DIR = os.path.join(BASE_DIR, "../banco_de_dados")
UPLOAD_FOLDER = os.path.join(DB_DIR, "img", "q_img")
ARQ_QUESTOES = os.path.join(DB_DIR, "questoes_concurso.xlsx")
ARQ_FLASHCARDS = os.path.join(DB_DIR, "flashcards.xlsx")

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# --- FUNÇÕES UTILITÁRIAS ---
def garantir_diretorio():
    if not os.path.exists(DB_DIR): os.makedirs(DB_DIR, exist_ok=True)


def limpar(texto):
    return str(texto).strip() if texto else ""

# --- NOVA FUNÇÃO: LIMPEZA PROFUNDA AO SALVAR ---
def normalizar_texto_para_banco(texto):
    if not texto: return ""
    txt = str(texto)

    # 1. Remove caracteres de retorno de carro do Windows (\r)
    txt = txt.replace('\r\n', '\n').replace('\r', '\n')

    # 2. Remove espaços em branco no final de cada linha
    txt = re.sub(r'[ \t]+\n', '\n', txt)

    # Remove espaços em branco no início de cada linha
    txt = re.sub(r'\n[ \t]+', '\n', txt)

    # 3. Colapsa 3 ou mais quebras de linha em apenas 1 (para manter parágrafo, mas sem buracos)
    txt = re.sub(r'\n{3,}', '\n\n', txt)

    return txt.strip()


def normalizar_para_comparacao(texto):
    if not texto: return ""
    texto_sem_tags = re.sub(r'<[^>]+>', '', str(texto))
    # Remove TUDO que não for letra ou número (espaços, tabs, quebras de linha, pontuação)
    # Isso gera uma string "pura" para comparação infalível
    return re.sub(r'[\W_]+', '', texto_sem_tags).lower().strip()


def sanitizar_texto(texto):
    if not texto: return ""
    # Remove hifens soltos de quebra de página
    texto = re.sub(r'-\s*\n\s*', '', texto)

    # Remove linhas isoladas de gabarito que possam ter sobrado (ex: em questões comentadas)
    # Mas cuidado para não remover partes do enunciado. O foco aqui é limpar "sujeira"
    texto = re.sub(r'\n\s*Gabarito:?\s*Letra\s*[A-E]\s*\n', '\n', texto, flags=re.IGNORECASE)

    linhas = [l.strip() for l in texto.split('\n') if l.strip()]
    if not linhas: return ""
    resultado = []
    for i in range(len(linhas)):
        atual = linhas[i]
        if i < len(linhas) - 1:
            proxima = linhas[i + 1]
            pontuacao_final = re.search(r'[.:?!;]$', atual)
            # Verifica se a próxima linha parece um novo bloco (começa com letra maiúscula ou número)
            comeca_novo_bloco = re.match(r'^(?:[A-Z"\'\(]|\d+\.|[a-e]\))', proxima)
            if not pontuacao_final and not comeca_novo_bloco:
                resultado.append(atual + " ")
            elif pontuacao_final and comeca_novo_bloco:
                resultado.append(atual + "\n")
            else:
                resultado.append(atual + " ")
        else:
            resultado.append(atual)
    return "".join(resultado).strip()


def gerar_assinatura(q):
    return (
        normalizar_para_comparacao(q.get('enunciado')),
        normalizar_para_comparacao(q.get('alt_a')),
        normalizar_para_comparacao(q.get('alt_b')),
        normalizar_para_comparacao(q.get('alt_c'))
    )


# --- RECONSTRUÇÃO DE TEXTO ---
def reconstruir_header_logico(texto):
    pattern = r"([A-Z\s\-\–]+)\n\s*((?:UESTÕES|ISTA).*)"

    def resolver_match(m):
        raw_letras = m.group(1)
        raw_palavras = m.group(2)
        matches_guia = list(re.finditer(r'([A-Z]|-)', raw_letras))
        palavras_quebradas = raw_palavras.split()
        stopwords = ["VERBAL", "TRAIÇOEIROS", "PARA", "COM", "DE", "DA", "DO", "DOS", "DAS", "EM", "QUE", "SE"]
        resultado_final = ""
        idx_p2 = 0
        for i, match in enumerate(matches_guia):
            token = match.group(1)
            termo_para_adicionar = token
            if token != '-':
                while idx_p2 < len(palavras_quebradas):
                    palavra_atual = palavras_quebradas[idx_p2]
                    if palavra_atual.upper().strip(".,:;") in stopwords:
                        resultado_final += palavra_atual + " "
                        idx_p2 += 1
                    else:
                        break
                if idx_p2 < len(palavras_quebradas):
                    termo_para_adicionar = token + palavras_quebradas[idx_p2]
                    idx_p2 += 1
            resultado_final += termo_para_adicionar
            if i < len(matches_guia) - 1:
                fim_atual = match.end()
                inicio_prox = matches_guia[i + 1].start()
                if inicio_prox > fim_atual:
                    resultado_final += " "
        if idx_p2 < len(palavras_quebradas):
            resultado_final += " " + " ".join(palavras_quebradas[idx_p2:])
        return "\n" + re.sub(r'\s+', ' ', resultado_final).strip() + "\n"

    return re.sub(pattern, resolver_match, texto)


def limpar_ruido(texto, disciplina=""):
    texto = reconstruir_header_logico(texto)
    # Normaliza a palavra GABARITO que pode vir espaçada ou quebrada
    texto = re.sub(r'G\s*\n?\s*A\s*B\s*A\s*R\s*I\s*T\s*O', 'Gabarito', texto, flags=re.IGNORECASE)

    patterns_to_remove = [
        r"PETROBRAS \(Nível Superior\) Português\s*\d*",
        r"www\.estrategiaconcursos\.com\.br\s*\d*",
        r".*Ricardo Aciole.*",
        r"^\s*\d+\s*$", # REMOVE LINHAS QUE SÃO APENAS NÚMEROS
        r"Equipe Português Estratégia Concursos, Felipe Luccas",
        r"Aula \d+",
        r"==\w+==",
        r"^\.\d+\.\.\)\.",
        r"10763321451",
    ]
    if disciplina == "Conhecimentos Específicos":
        patterns_to_remove.extend([
            r"PETROBRAS \(Engenharia de Produção\)",
            r"Conhecimentos Específicos",
            r"Daniel Almeida",
            r".*Felipe Canella.*"
        ])
    elif disciplina == "Inglês":
        patterns_to_remove.extend([
            r"^.*PETROBRAS \(Nível Superior\) Inglês.*$",
            r"^.*Ena Smith.*$",
            r"^.*Available at:.*$",
            r"^\d+\s*de\s*[A-Z][a-z]+\s*de\s*\d+",
        ])

    for pattern in patterns_to_remove:
        texto = re.sub(pattern, "", texto, flags=re.MULTILINE | re.IGNORECASE)
    texto = re.sub(r'\n{3,}', '\n\n', texto)
    return texto


def extrair_mapa_gabaritos_local(texto_bloco):

    mapa = {}
    # Procura por número + (ponto/traço opcional) + (LETRA opcional) + A-E
    # Ex: "1. A", "1. Letra A", "01 - A"
    padrao_tabela = r'\b(\d+)[\.\-\s]+\s*(?:[Ll][Ee][Tt][Rr][Aa])?\s*([A-E]|Certo|Errado|C|E)(?=[\s\d\.\-]|$)'
    matches = re.finditer(padrao_tabela, texto_bloco, re.IGNORECASE)
    for m in matches:
        val = m.group(2).upper()
        if val == "CERTO": val = "C"
        elif val == "ERRADO": val = "E"
        mapa[m.group(1)] = val
    return mapa


def parsear_questoes(texto_bruto, disciplina=""):
    texto_limpo = limpar_ruido(texto_bruto, disciplina)

    questoes = []

    if disciplina == "Português" or disciplina == "Conhecimentos Específicos":

        # Segmentação por Blocos Lógicos
        regex_divisao_blocos = re.compile(
            r'((?:QUESTÕES\s+COMENTADAS|LISTA\s+(?:DE|E)\s+QUESTÕES)(?:.|\n)+?)(?=(?:QUESTÕES\s+COMENTADAS|LISTA\s+(?:DE|E)\s+QUESTÕES)|$)',
            re.IGNORECASE)

        blocos = [m.group(1) for m in regex_divisao_blocos.finditer(texto_limpo)]

        if not blocos:
            blocos = [texto_limpo]

        assunto_atual = "Geral"

        for bloco in blocos:
            # Detecta o assunto do bloco pelo título
            match_titulo = re.match(r'((?:QUESTÕES\s+COMENTADAS|LISTA\s+(?:DE|E)\s+QUESTÕES).+?)(?:\n|$)', bloco,
                                re.IGNORECASE)

            if match_titulo:
                linha_completa = match_titulo.group(1).strip()
                idx_primeiro_hifen = linha_completa.find('-')
                idx_ultimo_hifen = linha_completa.rfind('-')
                if idx_primeiro_hifen != -1 and idx_ultimo_hifen != -1 and idx_primeiro_hifen < idx_ultimo_hifen:
                    assunto_raw = linha_completa[idx_primeiro_hifen + 1: idx_ultimo_hifen].strip()
                    assunto_atual = re.sub(r'Cesgranrio', '', assunto_raw.title(), flags=re.IGNORECASE).strip()
                elif "CORRELAÇÃO" in linha_completa.upper():
                    assunto_atual = "Correlação Verbal"
                elif "SINTÁTICAS" in linha_completa.upper():
                    assunto_atual = "Funções Sintáticas"
                elif "SEMÂNTICO" in linha_completa.upper():
                    assunto_atual = "Campo Semântico"
                elif "SINÔNIMO" in linha_completa.upper():
                    assunto_atual = "Sinônimos e Antônimos"
                elif "DENOTAÇÃO" in linha_completa.upper():
                    assunto_atual = "Denotação e Conotação"

            banca = "CESGRANRIO"
            instituicao = ""
            ano = "2025"

            # Extrai o mapa de respostas contido neste bloco (agora pega inline também)
            mapa_gabaritos_local = extrair_mapa_gabaritos_local(bloco)

            if disciplina == "Português":
                # Regex estrita para identificar início de questão
                pattern_questao = re.compile(
                    r'^\s*(\d+)\.\s*(?:\(?)\s*((?:\(|CESGRANRIO|FGV|CEBRASPE|FCC|VUNESP|INSTITUTO|BANCO|PETROBRAS|EQUIPE|[A-Z][a-zçãõâêô]+).+?)\s*(?:\)?)\s*$',
                    re.MULTILINE
                )
            elif disciplina == "Conhecimentos Específicos":
                # Sem ^ (início de linha) e sem $ (fim de linha). Pega inline.
                pattern_questao = re.compile(r'(?:^|\n)\s*(\d+)\s*[\.\-\)]\s*(\(.*?\))', re.MULTILINE)
            matches_questoes = list(pattern_questao.finditer(bloco))

            # --- Extração do Conteúdo do Texto de Apoio ---
            texto_apoio_bloco = ""

            for i, m in enumerate(matches_questoes):
                q_numero = m.group(1)
                q_meta = m.group(2)

                if disciplina == "Português":
                    # Filtro para evitar falsos positivos (como "1. Noções..." no índice)
                    if not re.search(r'^\(|CESGRANRIO|FGV|CEBRASPE|FCC|VUNESP|INSTITUTO|BANCO|PETROBRAS',
                                     q_meta.upper().strip()):
                        continue
                elif disciplina == "Conhecimentos Específicos":
                    if len(q_meta) < 3:
                        continue


                start_index = m.end()
                end_index = matches_questoes[i + 1].start() if i + 1 < len(matches_questoes) else len(bloco)

                q_conteudo_bruto = bloco[start_index:end_index]

                # Remover tabela de gabarito do final do texto da questão
                # Se encontrar "Gabarito 1." ou "Gabarito 1 ", corta o texto ali.
                # Isso evita que a tabela vá para a Alternativa E da última questão.
                q_conteudo_bruto = re.split(r'\n\s*Gabarito\s+1[\.\s]', q_conteudo_bruto, flags=re.IGNORECASE)[0]

                # INSERÇÃO: Detecção Universal de Certo/Errado
                tipo = "ME"
                if re.search(r'\(\s*\)\s*(?:Certo|Errado)|(?:Certo|Errado)\s*\(\s*\)|julgue\s+o\s+item|julgue\s+os\s+itens', q_conteudo_bruto, re.IGNORECASE):
                    tipo = "CE"

                # Processamento de metadados (Banca, Ano, etc)
                # CORREÇÃO: Busca o ano via regex (19xx ou 20xx) antes de quebrar a string
                match_ano = re.search(r'\b(19|20)\d{2}\b', q_meta)
                if match_ano:
                    ano = match_ano.group(0)

                # Remove o ano encontrado da string para limpar a área para Banca/Instituição
                meta_sem_ano = q_meta
                if match_ano:
                    meta_sem_ano = q_meta.replace(ano, "")

                meta_limpa = meta_sem_ano.replace("–", "/").replace("-", "/")

                # Removemos parênteses extras que podem sobrar após a limpeza
                partes_meta = [p.strip().replace('(', '').replace(')', '') for p in meta_limpa.split('/') if p.strip()]

                # Filtra strings vazias resultantes
                partes_meta = [p for p in partes_meta if p.strip()]

                if len(partes_meta) > 0:
                    banca_cand = partes_meta[0].replace('(', '')
                    if len(banca_cand) > 2: banca = banca_cand
                if len(partes_meta) > 1: instituicao = partes_meta[1].replace(')', '')

                # Busca Gabarito
                gabarito = ""
                # 1. Prioridade: Comentário local (questões comentadas)
                if disciplina == "Português":
                    gabarito_pattern_local = r'(?:Gabarito|Gab\.?|Letra|Correta)[:\s\.]+\s*([A-E])'
                else:
                    # 2. [A-E](?![a-z]): Pega a letra A-E SÓ SE não tiver letra minúscula depois (Evita o A de Alternativa).
                    gabarito_pattern_local = r'(?:Gabarito|Gab\.?|Letra|Correta)[:\s\.]+\s*(?:(?:Alternativa|Opção)\s+)?(?:[\"“\']\s*)?([A-Ea-e])(?:[\"”\']\.?)?(?![a-z])'
                matches_gab = list(re.finditer(gabarito_pattern_local, q_conteudo_bruto.strip(), re.IGNORECASE))
                if matches_gab:
                    gab_raw = matches_gab[-1].group(1).upper()
                    if gab_raw in ["CERTO", "C"]:
                        gabarito = "C"
                    elif gab_raw in ["ERRADO", "E"]:
                        gabarito = "E"
                    else:
                        gabarito = gab_raw

                # 2. Fallback: Mapa local (listas de questões)
                # Só usa se não achou no comentário E se não parece ter comentário no texto
                if not gabarito and q_numero in mapa_gabaritos_local:
                    if "Comentário" not in q_conteudo_bruto and "COMENTÁRIO" not in q_conteudo_bruto.upper():
                        gabarito = mapa_gabaritos_local[q_numero]

                partes_coment = re.split(r"(?i)Comentários?[:\s\-]*", q_conteudo_bruto, maxsplit=1)

                comentario_extraido = ""
                if len(partes_coment) > 1:
                    comentario_extraido = partes_coment[1].strip()
                    # Remove qualquer resquício de pontuação ou número de página no início
                    comentario_extraido = re.sub(r'^[:\-\s\d\.]+', '', comentario_extraido)

                # --- EXTRAÇÃO DO ENUNCIADO/ALTERNATIVAS (Corte anterior ao comentário) ---
                content_no_comments = partes_coment[0].strip()
                content_no_comments = re.sub(r'www\.estrategia.*', '', content_no_comments)


                # Separação Enunciado/Alternativas
                if tipo == "CE":
                    enunciado = re.sub(r'\(\s*\)\s*(?:Certo|Errado)|(?:Certo|Errado)\s*\(\s*\)', '', content_no_comments,
                                       flags=re.IGNORECASE)
                    enunciado = sanitizar_texto(enunciado)
                    alts = {"A": "", "B": "", "C": "", "D": "", "E": ""}
                else:
                    # --- CORREÇÃO PARA FORMATO (A), (B)... ---
                    if disciplina == "Conhecimentos Específicos" or disciplina == "Inglês":
                        content_no_comments = re.sub(r'(?:^|\s)\(([A-E])\)(?=\s)', r'\n\1)', content_no_comments)

                    parts_alt = re.split(r'\b([A-E])\)', content_no_comments, flags=re.IGNORECASE)
                    enunciado = sanitizar_texto(parts_alt[0].strip())
                    alts = {"A": "", "B": "", "C": "", "D": "", "E": ""}
                    if len(parts_alt) > 1:
                        for k in range(1, len(parts_alt), 2):
                            letra = parts_alt[k].upper()
                            if k + 1 < len(parts_alt):
                                alts[letra] = sanitizar_texto(parts_alt[k + 1].strip())

                if enunciado:
                    if (tipo == "ME" and (alts["A"] or alts["B"])) or (tipo == "CE"):
                        questoes.append({
                            "temp_id": str(uuid.uuid4()),
                            "banca": banca, "instituicao": instituicao, "ano": ano,
                            "assunto": assunto_atual, "enunciado": enunciado,
                            "alt_a": alts["A"], "alt_b": alts["B"], "alt_c": alts["C"], "alt_d": alts["D"],
                            "alt_e": alts["E"],
                            "gabarito": gabarito, "dificuldade": "Médio", "tipo": tipo, "imagem": "",
                            "comentarios": comentario_extraido
                        })



    elif disciplina == "Inglês":

        # Segmentação por Blocos Lógicos
        regex_divisao_blocos = re.compile(
            r'((?:QUESTÕES\s+COMENTADAS|LISTA\s+(?:DE|E)\s+QUESTÕES)(?:.|\n)+?)(?=(?:QUESTÕES\s+COMENTADAS|LISTA\s+(?:DE|E)\s+QUESTÕES)|$)',
            re.IGNORECASE)

        blocos = [m.group(1) for m in regex_divisao_blocos.finditer(texto_limpo)]

        if not blocos:
            blocos = [texto_limpo]

        assunto_atual = "Interpretação de Texto"

        for bloco in blocos:

            banca = "CESGRANRIO"
            instituicao = ""
            ano = "2025"

            # O regex busca: Qualquer coisa -> Hífen/Travessão -> (Instituição) -> Hífen/Travessão -> (Banca)
            match_meta_ing = re.search(r'.+?\s*[-–]\s*(.+?)\s*[-–]\s*(.+?)\s*(?:\n|$)', bloco[:600])
            if match_meta_ing:
                instituicao = match_meta_ing.group(1).strip()  # Grupo 1: BNDES
                banca = match_meta_ing.group(2).strip()  # Grupo 2: CESGRANRIO
            # Extrai o mapa de respostas contido neste bloco (agora pega inline também)
            mapa_gabaritos_local = extrair_mapa_gabaritos_local(bloco)

            # --- NOVO FILTRO DE UNICIDADE ---
            # Identifica todos os inícios (Número + Espaço + Letra Maiúscula)
            pattern_questao = re.compile(r'(?i)(?:^|\n|Gabarito[:\s]*[A-E])\s*(\d+)[\.\s\)]+\s*(?=[A-Z])', re.MULTILINE)
            todos_matches = list(pattern_questao.finditer(bloco))

            matches_questoes = []
            numeros_vistos = set()

            print(f"\n--- [DEBUG] INICIANDO CAPTURA NO BLOCO ---")

            for idx, m in enumerate(todos_matches):
                q_num = m.group(1)
                pos_match = m.start()

                # Define o fim da janela de busca: o início do próximo número encontrado ou o fim do bloco
                prox_pos_candidata = todos_matches[idx + 1].start() if idx + 1 < len(todos_matches) else len(bloco)
                janela_de_texto = bloco[m.end():prox_pos_candidata]

                # VALIDAÇÃO: Não importa se o enunciado é longo.
                # Se for uma QUESTÃO, as alternativas (A e B) DEVEM aparecer antes do próximo número.
                # Se for um NÚMERO DE LINHA (margem), as alternativas nunca aparecerão.
                tem_alts = re.search(r'\bA[\)\.]|\(A\)', janela_de_texto) and \
                           re.search(r'\bB[\)\.]|\(B\)', janela_de_texto)
                if q_num not in numeros_vistos and tem_alts:
                    print(f"[DEBUG - CAPTURA] Q{q_num} validada na pos {pos_match}")
                    matches_questoes.append(m)
                    numeros_vistos.add(q_num)
                else:
                    razao = "repetida" if q_num in numeros_vistos else "número de margem/sem alternativas"
                    print(f"[DEBUG - DESCARTADA] Ocorrência {q_num} na pos {pos_match} -> {razao}")

            matches_questoes.sort(key=lambda m: m.start())

            print(f"--- [DEBUG] INICIANDO PARSING INDIVIDUAL ---")
            for i, m in enumerate(matches_questoes):
                q_numero = m.group(1)
                start_index = m.end()
                end_index = matches_questoes[i + 1].start() if i + 1 < len(matches_questoes) else len(bloco)

                # O end_index define o limite da questão atual. Se estiver errado, a questão "engole" a próxima.
                if i + 1 < len(matches_questoes):
                    prox_match = matches_questoes[i + 1]
                    end_index = prox_match.start()
                    print(
                        f"[DEBUG - PROCESSANDO] Q{q_numero} -> Fim definido pelo início da Q{prox_match.group(1)} na posição {end_index}")
                else:
                    end_index = len(bloco)
                    print(f"[DEBUG - PROCESSANDO] Q{q_numero} -> Última questão do bloco. Fim na posição {end_index}")
                # 1. PEGA O BLOCO BRUTO (GIGANTE)
                q_conteudo_bruto = bloco[start_index:end_index]

                # 2. BUSCA O GABARITO (Independente do corte)
                gabarito = ""
                match_gab = re.search(r'(?i)GABARITO\s*:\s*([A-E])(?![a-z])', q_conteudo_bruto)
                if match_gab:
                    gabarito = match_gab.group(1).upper()
                elif q_numero in mapa_gabaritos_local:
                    gabarito = mapa_gabaritos_local[q_numero]

                # 3. IDENTIFICAÇÃO DO ENUNCIADO PARA CRIAR A ÂNCORA
                # Fazemos o split temporário apenas no bloco bruto
                temp_norm = re.sub(r'(?:^|\s)\(([A-E])\)(?=\s)', r'\n\1)', q_conteudo_bruto)
                parts_alt_temp = re.split(r'\b([A-E])\)', temp_norm, flags=re.IGNORECASE)

                enunciado_temp = parts_alt_temp[0].strip()
                limiar = 15
                # Pega os primeiros 15 caracteres do enunciado para ser a âncora de repetição
                ancora = enunciado_temp[:limiar] if len(enunciado_temp) >= limiar else enunciado_temp

                # 4. CORTE DA REPETIÇÃO (Limpando o vazamento para a Alt E)
                corpo_util = q_conteudo_bruto
                comentarios_extraidos = "" # Inicializa vazio

                if ancora:
                    # Padroniza espaços para a busca ser robusta contra quebras de linha do PDF
                    padrao_ancora = re.sub(r'\s+', r'\\s+', re.escape(ancora))
                    matches_ancora = list(re.finditer(padrao_ancora, q_conteudo_bruto, re.IGNORECASE))

                    # Se achou a repetição (segunda ocorrência), corta ali
                    if len(matches_ancora) > 1:
                        posicao_corte = matches_ancora[1].start()
                        corpo_util = q_conteudo_bruto[:posicao_corte].strip()

                        # TUDO após o corte vira comentário
                        comentarios_extraidos = q_conteudo_bruto[posicao_corte:].strip()
                    else:
                        # Fallback: Se não achou a repetição do texto, corta se o número da questão se repetir
                        # Isso evita o efeito dominó se a tradução começar com "11. De acordo..."
                        match_num = re.search(rf'\n\s*{q_numero}\s+', q_conteudo_bruto[limiar:])
                        if match_num:
                            posicao_corte = limiar + match_num.start()
                            corpo_util = q_conteudo_bruto[:posicao_corte].strip()
                            # TUDO após o corte vira comentário
                            comentarios_extraidos = q_conteudo_bruto[posicao_corte:].strip()

                # 5. SEPARAÇÃO FINAL DAS ALTERNATIVAS (Agora no corpo já cortado)
                content_final = re.sub(r'(?:^|\s)\(([A-E])\)(?=\s)', r'\n\1)', corpo_util)
                parts_alt = re.split(r'\b([A-E])\)', content_final, flags=re.IGNORECASE)

                enunciado = sanitizar_texto(parts_alt[0].strip())
                alts = {"A": "", "B": "", "C": "", "D": "", "E": ""}

                if len(parts_alt) > 1:
                    for k in range(1, len(parts_alt), 2):
                        letra = parts_alt[k].upper()
                        if k + 1 < len(parts_alt):
                            alts[letra] = sanitizar_texto(parts_alt[k + 1].strip())

                # 6. SALVAMENTO
                if enunciado and (alts["A"] or alts["B"]):
                    questoes.append({
                        "temp_id": str(uuid.uuid4()),
                        "banca": banca, "instituicao": instituicao, "ano": ano,
                        "assunto": assunto_atual, "enunciado": enunciado,
                        "alt_a": alts["A"], "alt_b": alts["B"], "alt_c": alts["C"],
                        "alt_d": alts["D"], "alt_e": alts["E"],
                        "gabarito": gabarito, "dificuldade": "Médio", "tipo": "ME",
                        "imagem": "", "comentarios": comentarios_extraidos
                    })

    print(f"--- [DEBUG] FIM DO PROCESSAMENTO ---\n")
    return questoes


def extrair_texto_pdf(caminho_arquivo):
    texto = ""
    with pdfplumber.open(caminho_arquivo) as pdf:
        for page in pdf.pages: texto += (page.extract_text() or "") + "\n"
    return texto


# --- CRUD BASE ---
def verificar_questoes():
    garantir_diretorio()
    if not os.path.exists(ARQ_QUESTOES):
        wb = Workbook();
        ws = wb.active;
        ws.title = "questoes"
        ws.append(
            ["id", "banca", "instituicao", "ano", "enunciado", "disciplina", "assunto", "dificuldade", "tipo", "alt_a",
             "alt_b", "alt_c", "alt_d", "alt_e", "gabarito", "respondidas", "acertos", "imagem", "comentarios",
             "texto_apoio"])
        wb.save(ARQ_QUESTOES)


def carregar_questoes():
    verificar_questoes();

    try:
        wb = load_workbook(ARQ_QUESTOES);
        ws = wb.active;
        dados = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None: continue

            img = row[17] if len(row) > 17 else ""
            coment = row[18] if len(row) > 18 else ""
            txt_apoio = row[19] if len(row) > 19 else ""

            dados.append({"id": row[0], "banca": row[1], "instituicao": row[2], "ano": row[3], "enunciado": row[4],
                          "disciplina": row[5], "assunto": row[6], "dificuldade": row[7], "tipo": row[8], "alt_a": row[9],
                          "alt_b": row[10], "alt_c": row[11], "alt_d": row[12], "alt_e": row[13], "gabarito": row[14],
                          "respondidas": row[15] or 0, "acertos": row[16] or 0, "imagem": img, "comentarios": coment,
                          "texto_apoio": txt_apoio})
        return dados
    except Exception as e:
        print(f"--- AVISO: Arquivo Excel ilegível ou corrompido ({e}) ---")
        return []

def salvar_questoes(dados):
    # 1. Tenta carregar o arquivo existente para PRESERVAR a aba 'textos'
    if os.path.exists(ARQ_QUESTOES):
        try:
            wb = load_workbook(ARQ_QUESTOES)
        except:
            wb = Workbook()  # Se o arquivo estiver corrompido, cria novo
    else:
        wb = Workbook()

    # 2. Gerencia a aba 'questoes'
    if "questoes" in wb.sheetnames:
        # Pega o índice para recriar na mesma posição visual
        idx = wb.sheetnames.index("questoes")
        wb.remove(wb["questoes"])
        ws = wb.create_sheet("questoes", idx)
    else:
        # Se não existe, cria como a primeira
        ws = wb.create_sheet("questoes", 0)

    # Remove aba padrão "Sheet" se ela foi criada automaticamente e está sobrando
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    ws.append(
        ["id", "banca", "instituicao", "ano", "enunciado", "disciplina", "assunto", "dificuldade", "tipo", "alt_a",
         "alt_b", "alt_c", "alt_d", "alt_e", "gabarito", "respondidas", "acertos", "imagem", "comentarios",
         "texto_apoio"])

    for i in dados:
        # APLICA A LIMPEZA AQUI: Garante que nada entra sujo no Excel
        enunciado_limpo = normalizar_texto_para_banco(i.get("enunciado", ""))
        alt_a = normalizar_texto_para_banco(i.get("alt_a", ""))
        alt_b = normalizar_texto_para_banco(i.get("alt_b", ""))
        alt_c = normalizar_texto_para_banco(i.get("alt_c", ""))
        alt_d = normalizar_texto_para_banco(i.get("alt_d", ""))
        alt_e = normalizar_texto_para_banco(i.get("alt_e", ""))

        coment_limpo = normalizar_texto_para_banco(i.get("comentarios", ""))

        ws.append(
            [i["id"], i.get("banca"), i.get("instituicao"), i.get("ano"),
             enunciado_limpo, i["disciplina"], i["assunto"],
             i["dificuldade"], i["tipo"],
             alt_a, alt_b, alt_c, alt_d, alt_e,
             i["gabarito"], i["respondidas"], i["acertos"], i.get("imagem", ""), coment_limpo,
             i.get("texto_apoio", "")])

    try:
        wb.save(ARQ_QUESTOES)
    except PermissionError:
        print("--- ERRO: Excel aberto. Não foi possível salvar as questões. ---")


# --- GERENCIAMENTO DE TEXTOS DE APOIO ---
def verificar_tabela_textos():
    garantir_diretorio()
    if not os.path.exists(ARQ_QUESTOES):
        return  # Se não existe o arquivo, a função verificar_questoes vai criar

    wb = load_workbook(ARQ_QUESTOES)
    if "textos" not in wb.sheetnames:
        ws = wb.create_sheet("textos")
        ws.append(["id", "titulo", "conteudo"])  # Cabeçalho
        wb.save(ARQ_QUESTOES)


def carregar_todos_textos():
    verificar_tabela_textos()
    wb = load_workbook(ARQ_QUESTOES)
    if "textos" not in wb.sheetnames: return []
    ws = wb["textos"]
    textos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            textos.append({"id": row[0], "titulo": row[1], "conteudo": row[2]})
    return textos


def salvar_novo_texto(novo_texto):
    try:
        verificar_tabela_textos()
        wb = load_workbook(ARQ_QUESTOES)
        ws = wb["textos"]

        ws.append([novo_texto["id"], novo_texto["titulo"], novo_texto["conteudo"]])
        wb.save(ARQ_QUESTOES)
        return True
    except PermissionError:
        print("--- ERRO CRÍTICO: Arquivo Excel aberto. Feche para salvar. ---")
        return False  # Retorna Falha
    except Exception as e:
        print(f"--- ERRO DESCONHECIDO: {e} ---")
        return False

def verificar_flashcards():
    garantir_diretorio()
    if not os.path.exists(ARQ_FLASHCARDS):
        wb = Workbook();
        ws = wb.active;
        ws.title = "flashcards"
        ws.append(["id", "disciplina", "assunto", "frente", "verso", "acertos", "erros"]);
        wb.save(ARQ_FLASHCARDS)


def carregar_flashcards():
    verificar_flashcards();
    wb = load_workbook(ARQ_FLASHCARDS);
    ws = wb.active;
    dados = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0]: dados.append(
            {"id": r[0], "disciplina": r[1], "assunto": r[2], "frente": r[3], "verso": r[4], "acertos": r[5] or 0,
             "erros": r[6] or 0})
    return dados


def salvar_flashcards_dados(dados):
    wb = Workbook();
    ws = wb.active;
    ws.append(["id", "disciplina", "assunto", "frente", "verso", "acertos", "erros"])
    for i in dados: ws.append(
        [i["id"], i["disciplina"], i["assunto"], i["frente"], i["verso"], i["acertos"], i["erros"]])
    wb.save(ARQ_FLASHCARDS)


def extrair_opcoes_do_banco():
    questoes = carregar_questoes()
    bancas = set()
    instituicoes = set()
    disciplinas = set()
    assuntos_map = {}

    for q in questoes:
        if q['banca']: bancas.add(str(q['banca']).strip().upper())
        if q['instituicao']: instituicoes.add(str(q['instituicao']).strip().upper())
        disc_raw = str(q['disciplina']).strip()
        disc_norm = disc_raw.title() if disc_raw else ""
        assunto_raw = str(q['assunto']).strip()

        if disc_norm:
            disciplinas.add(disc_norm)
            if disc_norm not in assuntos_map: assuntos_map[disc_norm] = set()
            if assunto_raw: assuntos_map[disc_norm].add(assunto_raw)

    assuntos_final = []
    for disc, lista_assuntos in assuntos_map.items():
        for a in sorted(list(lista_assuntos), key=str.lower):
            assuntos_final.append({'nome': a, 'disciplina': disc})

    if not disciplinas: disciplinas.add("Geral")
    if not bancas: bancas.add("BANCA PADRÃO")

    return {
        "bancas": sorted(list(bancas)),
        "instituicoes": sorted(list(instituicoes)),
        "disciplinas": sorted(list(disciplinas)),
        "assuntos": assuntos_final
    }


# --- ROTAS ---
@app.route('/img/q_img/<filename>')
def serve_image(filename): return send_from_directory(UPLOAD_FOLDER, filename)


@app.route("/questoes", methods=["GET"])
def get_q():
    dados = carregar_questoes()
    textos = carregar_todos_textos()
    mapa_textos = {t["id"]: t for t in textos}
    dados.reverse()  # Mais recentes primeiro

    # --- 1. FILTRAGEM (Server-Side) ---
    # Captura os parâmetros da URL
    texto = request.args.get('texto', '').lower()
    banca = request.args.get('banca', '')
    instituicao = request.args.get('instituicao', '')
    disciplina = request.args.get('disciplina', '')
    assunto = request.args.get('assunto', '')
    dificuldade = request.args.get('dificuldade', '')

    # Se houver algum filtro, aplicamos antes de paginar
    if any([texto, banca, instituicao, disciplina, assunto, dificuldade]):
        filtrados = []
        for q in dados:
            # Filtro Texto (Enunciado)
            if texto and texto not in (str(q.get('enunciado') or '')).lower(): continue

            # Filtros Exatos (Banca, Disciplina, etc)
            # Usamos 'or ""' para evitar erro se o campo for None
            if banca and str(q.get('banca') or '') != banca: continue
            if instituicao and str(q.get('instituicao') or '').upper() != instituicao.upper(): continue
            if disciplina and str(q.get('disciplina') or '') != disciplina: continue
            if assunto and str(q.get('assunto') or '') != assunto: continue
            if dificuldade and str(q.get('dificuldade') or '') != dificuldade: continue

            filtrados.append(q)
        dados = filtrados

    # Antes de paginar ou retornar, "hidrata" as questões com o texto real
    for q in dados:

        id_vinculo = q.get("texto_apoio")
        if id_vinculo and id_vinculo in mapa_textos:
            q["texto_titulo"] = mapa_textos[id_vinculo]["titulo"]
            q["texto_conteudo"] = mapa_textos[id_vinculo]["conteudo"]
        else:
            q["texto_titulo"] = ""
            q["texto_conteudo"] = ""

    # --- 2. PAGINAÇÃO ---
    page = request.args.get('page')
    if page:
        try:
            page = int(page)
            per_page = 50
            total_items = len(dados)
            total_pages = (total_items + per_page - 1) // per_page

            # Proteção: Se a página pedida for maior que o total, devolve a última
            if page > total_pages and total_pages > 0: page = total_pages
            if page < 1: page = 1

            start = (page - 1) * per_page
            end = start + per_page

            return jsonify({
                "items": dados[start:end],
                "total": total_items,
                "pagina_atual": page,
                "total_paginas": total_pages if total_pages > 0 else 1
            })
        except ValueError:
            pass

            # Retorno padrão (sem paginação, ex: para o modo Simulado)
    return jsonify(dados)


@app.route("/questoes", methods=["POST"])
def post_q():
    nova = {};
    arq = None
    if request.content_type.startswith('multipart'):
        nova = request.form.to_dict();
        arq = request.files.get('imagem_file')
    else:
        nova = request.json
    dados = carregar_questoes();
    if arq and arq.filename:
        ext = arq.filename.rsplit('.', 1)[1].lower();
        nome_img = f"{uuid.uuid4()}.{ext}";
        arq.save(os.path.join(UPLOAD_FOLDER, nome_img));
        nova["imagem"] = nome_img
    else:
        nova["imagem"] = ""
    sig = gerar_assinatura(nova)
    if any(gerar_assinatura(q) == sig for q in dados): return jsonify({"erro": "Duplicada"}), 409
    if not nova.get("id"): ids = sorted([int(q["id"]) for q in dados if str(q["id"]).isdigit()]); nova[
        "id"] = 1 if not ids else (ids[-1] + 1)
    nova.update({"respondidas": 0, "acertos": 0});
    dados.append(nova);
    salvar_questoes(dados)
    return jsonify({"mensagem": "Salvo", "id": nova["id"]}), 201


@app.route("/questoes", methods=["PUT"])
def put_q():
    load = {};
    arq = None
    if request.content_type and request.content_type.startswith('multipart'):
        load = request.form.to_dict();
        arq = request.files.get('imagem_file')
    else:
        load = request.json;
    if isinstance(load, list): salvar_questoes(load); return jsonify({"status": "ok"})
    dados = carregar_questoes()
    for i, q in enumerate(dados):
        if str(q["id"]) == str(load["id"]):
            img_antiga = q.get("imagem", "");
            q.update(load)
            if arq and arq.filename:
                ext = arq.filename.rsplit('.', 1)[1].lower();
                nome_img = f"{uuid.uuid4()}.{ext}";
                arq.save(os.path.join(UPLOAD_FOLDER, nome_img));
                q["imagem"] = nome_img
            else:
                if not load.get("imagem"): q["imagem"] = img_antiga
            dados[i] = q;
            salvar_questoes(dados);
            return jsonify({"status": "Atualizado"})
    return jsonify({"erro": "404"}), 404


@app.route("/check-duplicidade", methods=["POST"])
def check_dup():
    payload = request.json;
    sig = gerar_assinatura(payload)
    if not payload.get("enunciado"): return jsonify({"existe": False})
    dados = carregar_questoes();
    return jsonify({"existe": any(gerar_assinatura(q) == sig for q in dados)})


@app.route("/questoes/<string:id>", methods=["DELETE"])
def del_q(id):
    try:
        dados_atuais = carregar_questoes()
        novos_dados = [q for q in dados_atuais if str(q["id"]) != str(id)]

        # Se o tamanho for igual, não achou a questão (evita reescrever o arquivo à toa)
        if len(novos_dados) == len(dados_atuais):
            return jsonify({"status": "Questão não encontrada"}), 404

        salvar_questoes(novos_dados)
        return jsonify({"status": "Removido"})
    except PermissionError:
        return jsonify({"erro": "Arquivo Excel aberto. Feche para excluir."}), 500
    except Exception as e:
        print(f"Erro ao excluir: {e}")
        return jsonify({"erro": str(e)}), 500


@app.route("/flashcards", methods=["GET", "POST", "PUT"])
def handle_fc():
    try:
        # Se for apenas leitura
        if request.method == "GET":
            return jsonify(carregar_flashcards())

        # Carrega dados para operações de escrita
        dados = carregar_flashcards();
        load = request.json
        if request.method == "POST":
            # Gera ID novo se não vier
            if not load.get("id"):
                ids = sorted([int(f["id"]) for f in dados if str(f["id"]).isdigit()]);
                load["id"] = 1 if not ids else (ids[-1] + 1)

            # Inicializa contadores
            load.update({"acertos": 0, "erros": 0});
            dados.append(load);

            salvar_flashcards_dados(dados);
            return jsonify({"mensagem": "Salvo", "id": load["id"]}), 201

        if request.method == "PUT":
            found = False
            for i, f in enumerate(dados):
                if str(f["id"]) == str(load["id"]):
                    # Isso garante que campos não enviados (como acertos/erros) sejam mantidos
                    dados[i].update(load);

                    if "acertos" not in dados[i]: dados[i]["acertos"] = 0
                    if "erros" not in dados[i]: dados[i]["erros"] = 0

                    found = True
                    break

            if found:
                salvar_flashcards_dados(dados)
                return jsonify({"status": "Atualizado"})
            else:
                return jsonify({"erro": "404"}), 404

    except PermissionError:
        return jsonify({"erro": "O Excel de Flashcards está aberto. Feche-o."}), 500
    except Exception as e:
        print(f"Erro Flashcards: {e}")
        return jsonify({"erro": str(e)}), 500


@app.route("/flashcards/<string:id>", methods=["DELETE"])
def del_fc(id): dados = [f for f in carregar_flashcards() if str(f["id"]) != str(id)]; salvar_flashcards_dados(
    dados); return jsonify({"status": "Ok"})


@app.route("/upload-pdf", methods=["POST"])
def upload_pdf():
    f = request.files.get('file');
    disciplina = request.form.get('disciplina', '')

    if not f: return jsonify({"erro": "Sem arquivo"}), 400
    if not disciplina:
        return jsonify({"erro": "⚠️ Erro: Nenhuma disciplina selecionada."}), 400

    p = os.path.join(BASE_DIR, "temp.pdf");
    f.save(p)
    try:
        novas = parsear_questoes(extrair_texto_pdf(p), disciplina);
        banco = carregar_questoes();
        sigs = {gerar_assinatura(q) for q in banco}
        for n in novas: n['ja_cadastrada'] = gerar_assinatura(n) in sigs

        return jsonify(novas)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"erro": str(e)}), 500
    finally:
        if os.path.exists(p): os.remove(p)


@app.route("/opcoes-dinamicas", methods=["GET"])
def get_opcoes():
    return jsonify(extrair_opcoes_do_banco())


@app.route("/textos", methods=["GET", "POST"])
def handle_textos():
    if request.method == "GET":
        return jsonify(carregar_todos_textos())

    if request.method == "POST":
        data = request.json
        if not data.get("conteudo"): return jsonify({"erro": "Vazio"}), 400

        novo = {
            "id": str(uuid.uuid4()),
            "titulo": data.get("titulo", "Sem Título"),
            "conteudo": normalizar_texto_para_banco(data.get("conteudo"))
        }
        salvar_novo_texto(novo)
        return jsonify(novo), 201


if __name__ == "__main__": app.run(debug=True, port=5000)