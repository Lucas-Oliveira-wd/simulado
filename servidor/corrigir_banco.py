import os
import re
import shutil
from openpyxl import load_workbook

# --- CONFIGURAÇÕES ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_DIR = os.path.join(BASE_DIR, "../banco_de_dados")
ARQ_QUESTOES = os.path.join(DB_DIR, "questoes_concurso.xlsx")
ARQ_BACKUP = os.path.join(DB_DIR, "questoes_concurso_BACKUP.xlsx")


def limpar_texto_profundo(texto):
    if not texto or texto is None:
        return ""

    txt = str(texto)

    # 1. Remove caracteres de retorno de carro do Windows (\r)
    txt = txt.replace('\r', '')

    # 2. O PULO DO GATO: Substitui 2 ou mais quebras de linha (\n\n...) por APENAS UMA (\n)
    # Isso "colapsa" o texto, juntando os buracos, mas mantendo a quebra simples.
    txt = re.sub(r'\n{2,}', '\n', txt)

    # 3. Remove espaços em branco excessivos no início/fim de cada linha
    # Ex: "   Texto   \n" vira "Texto\n"
    linhas = [linha.strip() for linha in txt.split('\n')]
    txt = '\n'.join(linhas)

    return txt.strip()


def executar_limpeza():
    if not os.path.exists(ARQ_QUESTOES):
        print(f"ERRO: Arquivo não encontrado em {ARQ_QUESTOES}")
        return

    print("--- INICIANDO LIMPEZA DO BANCO DE DADOS ---")

    # 1. Cria um backup por segurança
    try:
        shutil.copy2(ARQ_QUESTOES, ARQ_BACKUP)
        print(f"1. Backup criado com sucesso: {ARQ_BACKUP}")
    except Exception as e:
        print(f"Erro ao criar backup: {e}")
        return

    # 2. Carrega a planilha
    try:
        wb = load_workbook(ARQ_QUESTOES)
        ws = wb.active
        print(f"2. Planilha carregada. Processando {ws.max_row} linhas...")
    except Exception as e:
        print(f"Erro ao abrir planilha: {e}")
        return

    # Índices das colunas (Baseado na ordem do seu código: A=1, B=2...)
    # Enunciado é a coluna 5 (E)
    # Alternativas são colunas 10, 11, 12, 13, 14 (J, K, L, M, N)
    colunas_para_limpar = [5, 10, 11, 12, 13, 14]

    contador_alteracoes = 0

    # Itera sobre as linhas (começando da 2 para pular o cabeçalho)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in colunas_para_limpar:
            celula = row[col_idx - 1]  # -1 porque lista é base 0
            valor_original = celula.value

            if valor_original:
                valor_limpo = limpar_texto_profundo(valor_original)

                # Só atualiza se houve mudança (para ganhar performance)
                if valor_original != valor_limpo:
                    celula.value = valor_limpo
                    contador_alteracoes += 1

    # 3. Salva o arquivo
    try:
        wb.save(ARQ_QUESTOES)
        print("------------------------------------------------")
        print(f"SUCESSO! O banco foi limpo e salvo.")
        print(f"Total de células corrigidas: {contador_alteracoes}")
        print("------------------------------------------------")
        print("Agora reinicie seu servidor Flask (coletar_diarios.py) para ver as mudanças.")
    except Exception as e:
        print(f"Erro ao salvar planilha: {e}")


if __name__ == "__main__":
    executar_limpeza()